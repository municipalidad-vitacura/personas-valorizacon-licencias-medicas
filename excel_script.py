from __future__ import annotations

from calendar import monthrange
from datetime import date, datetime, timedelta
from io import BytesIO
import json
import re
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode
from urllib.request import urlopen

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


AREAS_VALIDAS = {"Municipal", "Salud", "Educación"}
DOMINIO_POR_AREA = {"Municipal": 11, "Educación": 2, "Salud": 3}
EXTENSIONES_EXCEL = {".xlsx", ".xlsm", ".xltx", ".xltm"}
COLUMNAS_REQUERIDAS = ("rut", "fecha_inicio", "fecha_fin")
NOMBRE_HOJA_DETALLE = "detalle_mensual"
ENCABEZADOS_DETALLE = (
    "fila_origen",
    "id_licencia",
    "rut",
    "area",
    "periodo_inicio",
    "periodo_fin",
    "mes",
    "remuneracion",
    "dias_habiles",
    "valor_dia",
    "monto_mes",
)
MESES_ES = {
    1: "enero",
    2: "febrero",
    3: "marzo",
    4: "abril",
    5: "mayo",
    6: "junio",
    7: "julio",
    8: "agosto",
    9: "septiembre",
    10: "octubre",
    11: "noviembre",
    12: "diciembre",
}

FERIADOS = [
    "2023-01-01",
    "2023-01-02",
    "2023-04-07",
    "2023-04-08",
    "2023-05-01",
    "2023-05-07",
    "2023-05-21",
    "2023-06-21",
    "2023-06-26",
    "2023-07-16",
    "2023-08-15",
    "2023-09-18",
    "2023-09-19",
    "2023-10-09",
    "2023-10-27",
    "2023-11-01",
    "2023-12-08",
    "2023-12-17",
    "2023-12-25",
    "2024-01-01",
    "2024-03-29",
    "2024-03-30",
    "2024-05-01",
    "2024-05-21",
    "2024-06-09",
    "2024-06-20",
    "2024-06-29",
    "2024-07-16",
    "2024-08-15",
    "2024-09-18",
    "2024-09-19",
    "2024-09-20",
    "2024-10-12",
    "2024-10-27",
    "2024-10-31",
    "2024-11-01",
    "2024-11-24",
    "2024-12-08",
    "2024-12-25",
    "2025-01-01",
    "2025-04-18",
    "2025-04-19",
    "2025-05-01",
    "2025-05-21",
    "2025-06-20",
    "2025-06-29",
    "2025-07-16",
    "2025-08-15",
    "2025-09-18",
    "2025-09-19",
    "2025-10-12",
    "2025-10-31",
    "2025-11-01",
    "2025-11-16",
    "2025-12-08",
    "2025-12-14",
    "2025-12-25",
    "2026-01-01",
    "2026-04-03",
    "2026-04-04",
    "2026-05-01",
    "2026-05-21",
    "2026-06-21",
    "2026-06-29",
    "2026-07-16",
    "2026-08-15",
    "2026-09-18",
    "2026-09-19",
    "2026-10-12",
    "2026-10-31",
    "2026-11-01",
    "2026-12-08",
    "2026-12-25",
]
FERIADOS_CHILE = {date.fromisoformat(d) for d in FERIADOS}
VALIDADOR_RESPUESTA_KEYS = (
    "tipo_licencia",
    "dias",
    "fecha_desde",
    "fecha_hasta",
    "lugar_reposo",
    "entidad",
    "observacion",
)
FACTOR_PMG_MENSUALIZADO_FUNCIONARIO_NUEVO = 0.306
FACTOR_PMG_MENSUALIZADO_FUNCIONARIO_ANTIGUO = 0.226
COLUMNAS_SALIDA = ("monto_devolucion", "validador_licencia_valida") + tuple(
    f"validador_{k}" for k in VALIDADOR_RESPUESTA_KEYS
)
POSIBLES_KEYS_ITEMS = (
    "items",
    "detalle",
    "detalles",
    "detalle_liquidacion",
    "liquidacion_detalle",
)


class RRHHClient:
    BASE_URL = "http://appl.smc.cl/ws/wsjsonrrhh/usuario"

    def __init__(
        self,
        dominio_id: int,
        timeout: int = 20,
        *,
        anio_referencia: int | None = None,
    ) -> None:
        self.dominio_id = dominio_id
        self.timeout = timeout
        self.anio_referencia = anio_referencia if anio_referencia is not None else date.today().year
        self._cache_licencias: dict[str, list[dict[str, Any]]] = {}
        self._cache_proceso: dict[tuple[str, int, int], int | None] = {}
        self._cache_detalle: dict[tuple[str, int, int, int], dict[str, Any]] = {}
        self._cache_remuneracion: dict[tuple[str, int, int], float] = {}

    def obtener_licencias_medicas(self, rut: str) -> list[dict[str, Any]]:
        if rut in self._cache_licencias:
            return self._cache_licencias[rut]

        payload = self._get_json(
            "licencias_medicas.ashx",
            {"rut": rut, "dominio_id": self.dominio_id},
        )
        licencias = normalizar_lista_dicts(payload)
        self._cache_licencias[rut] = licencias
        return licencias

    def obtener_proceso_remuneraciones(self, rut: str, anio: int, mes: int) -> int | None:
        cache_key = (rut, anio, mes)
        if cache_key in self._cache_proceso:
            return self._cache_proceso[cache_key]

        payload = self._get_json(
            "liquidaciones_procesos.ashx",
            {"rut": rut, "anio": anio, "mes": mes, "dominio_id": self.dominio_id},
        )
        print(payload)
        procesos = normalizar_lista_dicts(payload)
        proceso_id: int | None = None
        for proceso in procesos:
            descripcion = str(proceso.get("descripcion", "")).strip().upper()
            if descripcion == "REMUNERACIONES":
                proceso_id = safe_int(proceso.get("nroliq"))
                break

        self._cache_proceso[cache_key] = proceso_id
        return proceso_id

    def obtener_detalle_liquidacion(
        self,
        rut: str,
        anio: int,
        mes: int,
        proceso_id: int,
    ) -> dict[str, Any]:
        cache_key = (rut, anio, mes, proceso_id)
        if cache_key in self._cache_detalle:
            return self._cache_detalle[cache_key]

        payload = self._get_json(
            "liquidaciones/detalle.ashx",
            {
                "rut": rut,
                "anio": anio,
                "proceso_id": proceso_id,
                "dominio_id": self.dominio_id,
                "mes": mes,
            },
        )
        detalle = normalizar_dict(payload)
        self._cache_detalle[cache_key] = detalle
        return detalle

    def obtener_remuneracion_mensual(self, rut: str, anio: int, mes: int) -> float:
        cache_key = (rut, anio, mes)
        if cache_key in self._cache_remuneracion:
            return self._cache_remuneracion[cache_key]

        proceso_id = self.obtener_proceso_remuneraciones(rut, anio, mes)
        if proceso_id is None:
            self._cache_remuneracion[cache_key] = 0.0
            return 0.0

        detalle = self.obtener_detalle_liquidacion(rut, anio, mes, proceso_id)
        total_haberes = safe_float(detalle.get("total_haberes"))
        valor_dl_3501 = 0.0
        for item in extraer_items_detalle(detalle):
            codigo = str(item.get("codigo", "")).strip().zfill(3)
            nombre = str(item.get("nombre", "")).strip().upper()
            if codigo == "004" and ("3501" in nombre or not nombre):
                valor_dl_3501 += safe_float(item.get("valor"))

        fecha_antiguedad = parsear_fecha_api(detalle.get("fecha_antiguedad"))
        factor_nuevo = FACTOR_PMG_MENSUALIZADO_FUNCIONARIO_NUEVO
        factor_antiguo = FACTOR_PMG_MENSUALIZADO_FUNCIONARIO_ANTIGUO
        es_nuevo = (
            fecha_antiguedad is not None
            and fecha_antiguedad.year == self.anio_referencia
        )
        factor = factor_nuevo if es_nuevo else factor_antiguo
        es_funcionario_dominio_11 = 1 if self.dominio_id == 11 else 0

        remuneracion = (
            total_haberes
            - valor_dl_3501
            + (factor * total_haberes * es_funcionario_dominio_11)
        )
        self._cache_remuneracion[cache_key] = remuneracion
        return remuneracion

    def _get_json(self, endpoint: str, params: dict[str, Any]) -> Any:
        query = urlencode(params)
        url = f"{self.BASE_URL}/{endpoint}?{query}"
        try:
            with urlopen(url, timeout=self.timeout) as response:
                payload = response.read().decode("utf-8")
        except HTTPError as exc:
            raise ValueError(
                f"Error HTTP consultando {endpoint}: {exc.code}"
            ) from exc
        except URLError as exc:
            raise ValueError(f"No se pudo conectar a {endpoint}.") from exc

        try:
            return json.loads(payload)
        except json.JSONDecodeError as exc:
            raise ValueError(f"JSON inválido en {endpoint}.") from exc


MODOS_ANIO_ARCHIVO = frozenset({"en_curso", "anteriores"})


def _colectar_anios_periodos_en_hoja(
    hoja_datos: Worksheet,
    indices: dict[str, int],
    max_col_original: int,
) -> set[int]:
    anios: set[int] = set()
    for fila_idx in range(2, hoja_datos.max_row + 1):
        fila_lista = [
            hoja_datos.cell(row=fila_idx, column=col_idx).value
            for col_idx in range(1, max_col_original + 1)
        ]
        if fila_vacia(fila_lista):
            continue
        raw_inicio = fila_lista[indices["fecha_inicio"] - 1]
        raw_fin = fila_lista[indices["fecha_fin"] - 1]
        try:
            fecha_inicio = parsear_fecha(raw_inicio)
            fecha_fin = parsear_fecha(raw_fin)
            if fecha_inicio > fecha_fin:
                continue
            anios.add(fecha_inicio.year)
            anios.add(fecha_fin.year)
        except ValueError:
            continue
    return anios


def _resolver_anio_referencia_archivo(
    modo: str,
    hoja_datos: Worksheet,
    indices: dict[str, int],
    max_col_original: int,
) -> int:
    if modo == "en_curso":
        return date.today().year
    anios = _colectar_anios_periodos_en_hoja(hoja_datos, indices, max_col_original)
    if anios:
        return max(anios)
    return date.today().year - 1


def modificar_archivo(
    contenido: bytes,
    extension: str,
    area: str,
    modo_anio_archivo: str = "en_curso",
) -> tuple[BytesIO, dict[str, Any]]:
    extension = extension.lower()
    if extension not in EXTENSIONES_EXCEL:
        raise ValueError("Formato no soportado para procesamiento de Excel.")

    if area not in AREAS_VALIDAS:
        raise ValueError("Área inválida. Usa Municipal, Salud o Educación.")

    if modo_anio_archivo not in MODOS_ANIO_ARCHIVO:
        raise ValueError(
            "Modo de año del archivo inválido. Usa año en curso o años anteriores."
        )

    dominio_id = DOMINIO_POR_AREA[area]
    workbook = cargar_workbook_desde_bytes(contenido, extension)
    hoja_datos = workbook.active
    indices = validar_encabezados(hoja_datos)
    max_col_original = hoja_datos.max_column
    anio_referencia = _resolver_anio_referencia_archivo(
        modo_anio_archivo,
        hoja_datos,
        indices,
        max_col_original,
    )
    cliente_rrhh = RRHHClient(
        dominio_id=dominio_id,
        anio_referencia=anio_referencia,
    )
    columnas_escritura = agregar_columnas_salida(hoja_datos)

    detalle_rows: list[tuple[Any, ...]] = []
    for fila_idx in range(2, hoja_datos.max_row + 1):
        fila_lista = [
            hoja_datos.cell(row=fila_idx, column=col_idx).value
            for col_idx in range(1, max_col_original + 1)
        ]
        if fila_vacia(fila_lista):
            continue

        id_licencia = valor_texto_por_header(
            fila_lista=fila_lista,
            indices=indices,
            header="id_licencia",
        )
        raw_rut = fila_lista[indices["rut"] - 1]
        raw_inicio = fila_lista[indices["fecha_inicio"] - 1]
        raw_fin = fila_lista[indices["fecha_fin"] - 1]

        try:
            rut = normalizar_rut(raw_rut)
            fecha_inicio = parsear_fecha(raw_inicio)
            fecha_fin = parsear_fecha(raw_fin)
            if fecha_inicio > fecha_fin:
                raise ValueError("fecha_inicio es mayor que fecha_fin.")
        except ValueError:
            escribir_salida_fila(
                hoja_datos,
                fila_idx,
                columnas_escritura,
                0.0,
                validador_vacio(False),
            )
            detalle_rows.append(
                (
                    fila_idx,
                    id_licencia,
                    valor_texto(fila_lista, indices["rut"]),
                    area,
                    formatear_fecha_salida(raw_inicio),
                    formatear_fecha_salida(raw_fin),
                    "error",
                    0.0,
                    0,
                    0.0,
                    0.0,
                )
            )
            continue

        validador_datos = validador_vacio(False)
        try:
            validador_datos = validar_licencia_medica(
                cliente_rrhh,
                rut,
                fecha_inicio,
                fecha_fin,
            )
        except Exception:  # noqa: BLE001
            validador_datos = validador_vacio(False)

        monto_devolucion = 0.0
        for inicio_tramo, fin_tramo in desglosar_periodo_mensual(fecha_inicio, fecha_fin):
            dias_habiles = calcular_dias_habiles_chile(inicio_tramo, fin_tramo)
            remuneracion = 0.0
            try:
                remuneracion = obtener_remuneracion(
                    cliente_rrhh,
                    rut,
                    inicio_tramo.year,
                    inicio_tramo.month,
                )
            except Exception:  # noqa: BLE001
                remuneracion = 0.0

            valor_dia = obtener_valor_dia(remuneracion)
            monto_mes = valor_dia * dias_habiles
            monto_devolucion += monto_mes

            detalle_rows.append(
                (
                    fila_idx,
                    id_licencia,
                    rut,
                    area,
                    inicio_tramo,
                    fin_tramo,
                    nombre_mes(inicio_tramo.month),
                    remuneracion,
                    dias_habiles,
                    valor_dia,
                    monto_mes,
                )
            )

        escribir_salida_fila(
            hoja_datos,
            fila_idx,
            columnas_escritura,
            monto_devolucion,
            validador_datos,
        )

    escribir_hoja_detalle(workbook, detalle_rows)
    salida = BytesIO()
    workbook.save(salida)
    salida.seek(0)
    meta: dict[str, Any] = {
        "modo_anio_archivo": modo_anio_archivo,
        "anio_referencia": anio_referencia,
    }
    return salida, meta


def validar_licencia_medica(
    cliente_rrhh: RRHHClient,
    rut: str,
    fecha_inicio: date,
    fecha_fin: date,
) -> dict[str, Any]:
    licencias = cliente_rrhh.obtener_licencias_medicas(rut)
    if not licencias:
        return validador_vacio(False)

    licencia_encontrada: dict[str, Any] | None = None
    for licencia in licencias:
        fecha_desde = parsear_fecha_api(licencia.get("fecha_desde"))
        fecha_hasta = parsear_fecha_api(licencia.get("fecha_hasta"))
        if fecha_desde == fecha_inicio and fecha_hasta == fecha_fin:
            licencia_encontrada = licencia
            break

    if licencia_encontrada is None:
        return validador_vacio(False)

    fecha_desde = parsear_fecha_api(licencia_encontrada.get("fecha_desde"))
    fecha_hasta = parsear_fecha_api(licencia_encontrada.get("fecha_hasta"))
    licencia_valida = fecha_desde == fecha_inicio and fecha_hasta == fecha_fin

    datos = validador_vacio(licencia_valida)
    for key in VALIDADOR_RESPUESTA_KEYS:
        value = licencia_encontrada.get(key)
        if key in {"fecha_desde", "fecha_hasta"}:
            value = parsear_fecha_api(value) or value
        datos[f"validador_{key}"] = value
    return datos


def cargar_workbook_desde_bytes(contenido: bytes, extension: str) -> Workbook:
    if not contenido:
        raise ValueError("El archivo está vacío.")

    keep_vba = extension in {".xlsm", ".xltm"}
    try:
        return load_workbook(filename=BytesIO(contenido), keep_vba=keep_vba)
    except Exception as exc:  # noqa: BLE001
        raise ValueError("No se pudo leer el archivo Excel.") from exc


def validar_encabezados(hoja: Worksheet) -> dict[str, int]:
    headers: dict[str, int] = {}
    for col_idx in range(1, hoja.max_column + 1):
        raw = hoja.cell(row=1, column=col_idx).value
        if raw is None:
            continue
        header = str(raw).strip()
        if header in COLUMNAS_REQUERIDAS:
            headers[header] = col_idx

    faltantes = [col for col in COLUMNAS_REQUERIDAS if col not in headers]
    if faltantes:
        requeridas = ", ".join(COLUMNAS_REQUERIDAS)
        faltantes_txt = ", ".join(faltantes)
        raise ValueError(
            f"Faltan columnas requeridas: {faltantes_txt}. "
            f"Se requieren exactamente: {requeridas}."
        )
    return headers


def agregar_columnas_salida(hoja: Worksheet) -> dict[str, int]:
    columnas: dict[str, int] = {}
    col = hoja.max_column + 1
    for nombre_columna in COLUMNAS_SALIDA:
        hoja.cell(row=1, column=col, value=nombre_columna)
        columnas[nombre_columna] = col
        col += 1
    return columnas


def escribir_salida_fila(
    hoja: Worksheet,
    fila: int,
    columnas: dict[str, int],
    monto_devolucion: float,
    validador_datos: dict[str, Any],
) -> None:
    hoja.cell(row=fila, column=columnas["monto_devolucion"], value=monto_devolucion)
    hoja.cell(
        row=fila,
        column=columnas["validador_licencia_valida"],
        value="SI" if validador_datos["validador_licencia_valida"] else "NO",
    )
    for key in VALIDADOR_RESPUESTA_KEYS:
        nombre = f"validador_{key}"
        hoja.cell(row=fila, column=columnas[nombre], value=validador_datos.get(nombre))


def validador_vacio(licencia_valida: bool) -> dict[str, Any]:
    salida = {"validador_licencia_valida": licencia_valida}
    for key in VALIDADOR_RESPUESTA_KEYS:
        salida[f"validador_{key}"] = None
    return salida


def fila_vacia(fila: list[Any]) -> bool:
    return all(celda is None or str(celda).strip() == "" for celda in fila)


def valor_texto(fila: list[Any], index_columna: int) -> str:
    raw = fila[index_columna - 1]
    if raw is None:
        return ""
    return str(raw).strip()


def valor_texto_por_header(
    fila_lista: list[Any],
    indices: dict[str, int],
    header: str,
) -> str:
    index = indices.get(header)
    if index is None:
        return ""
    return valor_texto(fila_lista, index)


def normalizar_rut(raw_rut: Any) -> str:
    texto = str(raw_rut).strip()
    if not texto:
        raise ValueError("RUT vacío.")

    texto = texto.replace(".", "").replace(" ", "")
    if "-" in texto:
        texto = texto.split("-", 1)[0]
    else:
        texto = re.sub(r"[^0-9A-Za-z]", "", texto)
        if len(texto) > 8:
            texto = texto[:-1]

    if not texto.isdigit():
        raise ValueError("RUT inválido. Debe quedar sin DV, sin puntos y sin guion.")
    return texto


def parsear_fecha(valor: Any) -> date:
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    if isinstance(valor, str):
        limpio = valor.strip()
        if not limpio:
            raise ValueError("Fecha vacía.")
        try:
            return datetime.strptime(limpio, "%d-%m-%y").date()
        except ValueError as exc:
            raise ValueError("Formato de fecha inválido, usa dd-mm-yy.") from exc
    raise ValueError("Tipo de dato de fecha no soportado.")


def parsear_fecha_api(valor: Any) -> date | None:
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    if valor is None:
        return None
    texto = str(valor).strip()
    if not texto:
        return None

    formatos = ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d", "%d-%m-%Y", "%d-%m-%y")
    for fmt in formatos:
        try:
            return datetime.strptime(texto, fmt).date()
        except ValueError:
            continue

    if "T" in texto:
        try:
            return datetime.fromisoformat(texto).date()
        except ValueError:
            return None
    return None


def formatear_fecha_salida(valor: Any) -> Any:
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    if isinstance(valor, str):
        return valor.strip()
    return valor


def desglosar_periodo_mensual(inicio: date, fin: date) -> list[tuple[date, date]]:
    if inicio > fin:
        raise ValueError("El inicio no puede ser mayor al fin.")

    tramos: list[tuple[date, date]] = []
    cursor = inicio
    while cursor <= fin:
        ultimo_dia_mes = monthrange(cursor.year, cursor.month)[1]
        fin_mes = date(cursor.year, cursor.month, ultimo_dia_mes)
        tramo_fin = min(fin, fin_mes)
        tramos.append((cursor, tramo_fin))
        cursor = tramo_fin + timedelta(days=1)
    return tramos


def calcular_dias_habiles_chile(inicio: date, fin: date) -> int:
    if inicio > fin:
        raise ValueError("El inicio no puede ser mayor al fin.")

    total = 0
    cursor = inicio
    while cursor <= fin:
        if cursor.weekday() < 5 and cursor not in FERIADOS_CHILE:
            total += 1
        cursor += timedelta(days=1)
    return total


def obtener_remuneracion(cliente_rrhh: RRHHClient, rut: str, anio: int, mes: int) -> float:
    return cliente_rrhh.obtener_remuneracion_mensual(rut=rut, anio=anio, mes=mes)


def obtener_valor_dia(remuneracion: float) -> float:
    if remuneracion == 0:
        return 0.0
    return remuneracion / 30


def nombre_mes(numero_mes: int) -> str:
    return MESES_ES[numero_mes]


def escribir_hoja_detalle(workbook: Workbook, detalle_rows: list[tuple[Any, ...]]) -> None:
    if NOMBRE_HOJA_DETALLE in workbook.sheetnames:
        workbook.remove(workbook[NOMBRE_HOJA_DETALLE])

    hoja_detalle = workbook.create_sheet(title=NOMBRE_HOJA_DETALLE)
    hoja_detalle.append(list(ENCABEZADOS_DETALLE))
    for row in detalle_rows:
        hoja_detalle.append(list(row))


def extraer_items_detalle(detalle: dict[str, Any]) -> list[dict[str, Any]]:
    for key in POSIBLES_KEYS_ITEMS:
        raw = detalle.get(key)
        if isinstance(raw, list):
            return [item for item in raw if isinstance(item, dict)]

    for value in detalle.values():
        if isinstance(value, list) and all(isinstance(item, dict) for item in value):
            if value and ("codigo" in value[0] or "valor" in value[0]):
                return value
    return []


def normalizar_lista_dicts(payload: Any) -> list[dict[str, Any]]:
    if isinstance(payload, list):
        return [item for item in payload if isinstance(item, dict)]

    if isinstance(payload, dict):
        for value in payload.values():
            if isinstance(value, list) and all(isinstance(item, dict) for item in value):
                return value
        return [payload]

    return []


def normalizar_dict(payload: Any) -> dict[str, Any]:
    if isinstance(payload, dict):
        return payload
    if isinstance(payload, list):
        for item in payload:
            if isinstance(item, dict):
                return item
    return {}


def safe_float(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    texto = str(value).strip()
    if not texto:
        return 0.0

    texto = texto.replace(" ", "")
    if "," in texto and "." in texto:
        if texto.rfind(",") > texto.rfind("."):
            texto = texto.replace(".", "").replace(",", ".")
        else:
            texto = texto.replace(",", "")
    elif "," in texto:
        texto = texto.replace(",", ".")

    try:
        return float(texto)
    except ValueError:
        return 0.0


def safe_int(value: Any) -> int:
    return int(round(safe_float(value)))
