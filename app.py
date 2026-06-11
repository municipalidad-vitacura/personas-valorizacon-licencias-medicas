from __future__ import annotations

import os
from io import BytesIO
from pathlib import Path
from typing import Any
from uuid import uuid4

from flask import Flask, flash, make_response, redirect, render_template, request, send_file
from openpyxl import load_workbook
from werkzeug.exceptions import RequestEntityTooLarge
from werkzeug.utils import secure_filename

from excel_script import modificar_archivo


ALLOWED_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}
MIME_BY_EXTENSION = {
    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ".xlsm": "application/vnd.ms-excel.sheet.macroEnabled.12",
    ".xltx": "application/vnd.openxmlformats-officedocument.spreadsheetml.template",
    ".xltm": "application/vnd.ms-excel.template.macroEnabled.12",
}
AREAS_VALIDAS = {"Municipal", "Salud", "Educación"}
MODOS_ANIO_ARCHIVO = frozenset({"en_curso", "anteriores"})
ETIQUETA_ANIO_ARCHIVO = {
    "en_curso": "Año en curso",
    "anteriores": "Años anteriores",
}
NOMBRE_HOJA_DETALLE = "detalle_mensual"
ENCABEZADOS_DETALLE = (
    "fila_origen",
    "id_licencia",
    "rut",
    "area",
    "periodo_inicio",
    "periodo_fin",
    "mes",
    "remuneracion",
    "dias_habiles",
    "valor_dia",
    "monto_mes",
)
SALIDAS_PROCESADAS: dict[str, dict[str, Any]] = {}

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 20 * 1024 * 1024  # 20 MB
app.secret_key = os.getenv("FLASK_SECRET_KEY", "dev-secret-key")


def extension_permitida(filename: str) -> bool:
    return Path(filename).suffix.lower() in ALLOWED_EXTENSIONS


def fila_vacia(fila: list[Any]) -> bool:
    return all(valor is None or str(valor).strip() == "" for valor in fila)


def formatear_valor(valor: Any) -> Any:
    if valor is None:
        return ""
    if hasattr(valor, "isoformat"):
        return valor.isoformat()
    return valor


def _obtener_indices_encabezado(hoja) -> dict[str, int]:
    indices: dict[str, int] = {}
    for col_idx in range(1, hoja.max_column + 1):
        raw = hoja.cell(row=1, column=col_idx).value
        if raw is None:
            continue
        indices[str(raw).strip()] = col_idx
    return indices


def construir_detalle_respuesta(contenido_excel: bytes) -> dict[str, Any] | None:
    workbook = load_workbook(filename=BytesIO(contenido_excel), data_only=True)
    hoja_datos = workbook.active
    indices = _obtener_indices_encabezado(hoja_datos)

    filas_utiles: list[tuple[int, list[Any]]] = []
    for fila_idx in range(2, hoja_datos.max_row + 1):
        fila_lista = [
            hoja_datos.cell(row=fila_idx, column=col_idx).value
            for col_idx in range(1, hoja_datos.max_column + 1)
        ]
        if fila_vacia(fila_lista):
            continue
        filas_utiles.append((fila_idx, fila_lista))

    if len(filas_utiles) != 1:
        return None

    fila_excel, fila_lista = filas_utiles[0]

    def valor_columna(nombre: str) -> Any:
        col_idx = indices.get(nombre)
        if col_idx is None:
            return ""
        return formatear_valor(fila_lista[col_idx - 1])

    resumen = {
        "fila_origen": fila_excel,
        "rut": valor_columna("rut"),
        "fecha_inicio": valor_columna("fecha_inicio"),
        "fecha_fin": valor_columna("fecha_fin"),
        "monto_devolucion": valor_columna("monto_devolucion"),
        "validador_licencia_valida": valor_columna("validador_licencia_valida"),
        "validador_tipo_licencia": valor_columna("validador_tipo_licencia"),
        "validador_dias": valor_columna("validador_dias"),
        "validador_fecha_desde": valor_columna("validador_fecha_desde"),
        "validador_fecha_hasta": valor_columna("validador_fecha_hasta"),
        "validador_lugar_reposo": valor_columna("validador_lugar_reposo"),
        "validador_entidad": valor_columna("validador_entidad"),
        "validador_observacion": valor_columna("validador_observacion"),
    }

    detalle_mensual: list[dict[str, Any]] = []
    if NOMBRE_HOJA_DETALLE in workbook.sheetnames:
        hoja_detalle = workbook[NOMBRE_HOJA_DETALLE]
        indices_detalle = _obtener_indices_encabezado(hoja_detalle)
        for fila_idx in range(2, hoja_detalle.max_row + 1):
            fila_origen = hoja_detalle.cell(
                row=fila_idx, column=indices_detalle.get("fila_origen", 0)
            ).value
            if fila_origen != fila_excel:
                continue

            item: dict[str, Any] = {}
            for nombre in ENCABEZADOS_DETALLE:
                col_idx = indices_detalle.get(nombre)
                if col_idx is None:
                    item[nombre] = ""
                    continue
                item[nombre] = formatear_valor(
                    hoja_detalle.cell(row=fila_idx, column=col_idx).value
                )
            detalle_mensual.append(item)

    return {
        "resumen": resumen,
        "detalle_mensual": detalle_mensual,
    }


@app.get("/")
def index():
    response = make_response(render_template("index.html"))
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    return response


@app.errorhandler(RequestEntityTooLarge)
def archivo_muy_grande(_error):
    flash("El archivo supera el tamano maximo permitido (20 MB).")
    return redirect("/")


@app.post("/procesar")
def procesar_excel():
    archivo = request.files.get("archivo")
    area = request.form.get("area", "")
    modo_anio_archivo = request.form.get("anio_archivo", "")

    if archivo is None or archivo.filename == "":
        flash("Debes seleccionar un archivo Excel.")
        return redirect("/")

    nombre_seguro = secure_filename(archivo.filename)
    extension = Path(nombre_seguro).suffix.lower()

    if modo_anio_archivo not in MODOS_ANIO_ARCHIVO:
        flash("Debes indicar si el archivo corresponde al año en curso o a años anteriores.")
        return redirect("/")

    if area not in AREAS_VALIDAS:
        flash("Debes elegir un área válida: Municipal, Salud o Educación.")
        return redirect("/")

    if not extension_permitida(nombre_seguro):
        flash("Formato no soportado. Usa: .xlsx, .xlsm, .xltx o .xltm.")
        return redirect("/")

    contenido = archivo.read()
    if not contenido:
        flash("El archivo está vacío.")
        return redirect("/")

    try:
        salida_io, meta_anio = modificar_archivo(
            contenido,
            extension,
            area,
            modo_anio_archivo=modo_anio_archivo,
        )
    except ValueError as exc:
        flash(str(exc))
        return redirect("/")

    base = Path(nombre_seguro).stem
    nombre_salida = f"{base}_modificado{extension}"
    salida_bytes = salida_io.getvalue()
    detalle = construir_detalle_respuesta(salida_bytes)
    meta_anio_vista = {
        "modo_codigo": meta_anio["modo_anio_archivo"],
        "modo_etiqueta": ETIQUETA_ANIO_ARCHIVO[meta_anio["modo_anio_archivo"]],
        "anio_referencia": meta_anio["anio_referencia"],
    }

    if detalle is not None:
        token_descarga = str(uuid4())
        SALIDAS_PROCESADAS[token_descarga] = {
            "contenido": salida_bytes,
            "nombre": nombre_salida,
            "mimetype": MIME_BY_EXTENSION[extension],
        }
        return render_template(
            "detalle.html",
            detalle=detalle,
            token_descarga=token_descarga,
            meta_anio=meta_anio_vista,
        )

    return send_file(
        BytesIO(salida_bytes),
        as_attachment=True,
        download_name=nombre_salida,
        mimetype=MIME_BY_EXTENSION[extension],
    )


@app.get("/descargar/<token>")
def descargar_procesado(token: str):
    payload = SALIDAS_PROCESADAS.pop(token, None)
    if payload is None:
        flash("La descarga expiró. Vuelve a procesar el archivo.")
        return redirect("/")

    return send_file(
        BytesIO(payload["contenido"]),
        as_attachment=True,
        download_name=payload["nombre"],
        mimetype=payload["mimetype"],
    )


if __name__ == "__main__":
    debug = os.getenv("FLASK_DEBUG", "1") == "1"
    app.run(host="0.0.0.0", port=5000, debug=debug)
