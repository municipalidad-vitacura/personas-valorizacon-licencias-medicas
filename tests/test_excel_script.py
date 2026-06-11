from __future__ import annotations

import unittest
from datetime import date
from io import BytesIO
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from excel_script import (
    calcular_dias_habiles_chile,
    desglosar_periodo_mensual,
    modificar_archivo,
)


class ExcelScriptTests(unittest.TestCase):
    def workbook_bytes(self, headers: list[str], rows: list[list[object]]) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        for row in rows:
            ws.append(row)

        out = BytesIO()
        wb.save(out)
        return out.getvalue()

    def test_desglosar_periodo_mensual(self) -> None:
        tramos = desglosar_periodo_mensual(date(2026, 1, 10), date(2026, 3, 24))
        self.assertEqual(
            tramos,
            [
                (date(2026, 1, 10), date(2026, 1, 31)),
                (date(2026, 2, 1), date(2026, 2, 28)),
                (date(2026, 3, 1), date(2026, 3, 24)),
            ],
        )

    def test_calcular_dias_habiles_chile_ejemplo(self) -> None:
        enero = calcular_dias_habiles_chile(date(2026, 1, 10), date(2026, 1, 31))
        febrero = calcular_dias_habiles_chile(date(2026, 2, 1), date(2026, 2, 28))
        marzo = calcular_dias_habiles_chile(date(2026, 3, 1), date(2026, 3, 24))
        self.assertEqual((enero, febrero, marzo), (15, 20, 17))

    @patch("excel_script.RRHHClient.obtener_remuneracion_mensual", return_value=300000.0)
    @patch(
        "excel_script.RRHHClient.obtener_licencias_medicas",
        return_value=[
            {
                "tipo_licencia": "Enfermedad Comun",
                "dias": 75,
                "fecha_desde": "2026-01-10T00:00:00",
                "fecha_hasta": "2026-03-24T00:00:00",
                "lugar_reposo": "DOMICILIO",
                "entidad": "CRUZ BLANCA",
                "observacion": "PAGADA",
            }
        ],
    )
    def test_modificar_archivo_agrega_monto_y_detalle(
        self,
        _mock_licencias,
        _mock_remuneracion,
    ) -> None:
        contenido = self.workbook_bytes(
            headers=["rut", "fecha_inicio", "fecha_fin"],
            rows=[["19.829.424-1", "10-01-26", "24-03-26"]],
        )

        salida, _meta = modificar_archivo(contenido, ".xlsx", "Municipal")
        result = load_workbook(filename=BytesIO(salida.getvalue()))
        ws = result.active

        self.assertEqual(ws.cell(row=1, column=4).value, "monto_devolucion")
        self.assertEqual(ws.cell(row=2, column=4).value, 520000.0)
        self.assertEqual(ws.cell(row=2, column=5).value, "SI")
        self.assertEqual(ws.cell(row=2, column=6).value, "Enfermedad Comun")
        self.assertEqual(ws.cell(row=2, column=10).value, "DOMICILIO")
        self.assertIn("detalle_mensual", result.sheetnames)

        detalle = result["detalle_mensual"]
        self.assertEqual(detalle.max_row, 4)  # header + 3 tramos
        self.assertEqual(detalle.cell(row=2, column=7).value, "enero")
        self.assertEqual(detalle.cell(row=3, column=7).value, "febrero")
        self.assertEqual(detalle.cell(row=4, column=7).value, "marzo")
        self.assertEqual(detalle.cell(row=2, column=8).value, 300000.0)
        self.assertEqual(detalle.cell(row=2, column=9).value, 15)
        self.assertEqual(detalle.cell(row=3, column=9).value, 20)
        self.assertEqual(detalle.cell(row=4, column=9).value, 17)

    def test_modificar_archivo_falta_columna_requerida(self) -> None:
        contenido = self.workbook_bytes(
            headers=["fecha_inicio", "fecha_fin"],
            rows=[["10-01-26", "24-03-26"]],
        )

        with self.assertRaises(ValueError):
            modificar_archivo(contenido, ".xlsx", "Municipal")

    def test_modificar_archivo_fecha_invalida(self) -> None:
        contenido = self.workbook_bytes(
            headers=["rut", "fecha_inicio", "fecha_fin"],
            rows=[["19829424", "31-02-26", "24-03-26"]],
        )

        salida, _meta = modificar_archivo(contenido, ".xlsx", "Salud")
        result = load_workbook(filename=BytesIO(salida.getvalue()))
        ws = result.active
        detalle = result["detalle_mensual"]

        self.assertEqual(ws.cell(row=2, column=4).value, 0)
        self.assertEqual(ws.cell(row=2, column=5).value, "NO")
        self.assertEqual(detalle.cell(row=2, column=7).value, "error")
        self.assertEqual(detalle.cell(row=2, column=11).value, 0)

    @patch("excel_script.RRHHClient.obtener_remuneracion_mensual", return_value=300000.0)
    @patch(
        "excel_script.RRHHClient.obtener_licencias_medicas",
        return_value=[
            {
                "tipo_licencia": "Enfermedad Comun",
                "dias": 5,
                "fecha_desde": "2026-01-20T00:00:00",
                "fecha_hasta": "2026-01-25T00:00:00",
                "lugar_reposo": "DOMICILIO",
                "entidad": "CRUZ BLANCA",
                "observacion": "PAGADA",
            }
        ],
    )
    def test_modificar_archivo_licencia_sin_match_fechas(
        self,
        _mock_licencias,
        _mock_remuneracion,
    ) -> None:
        contenido = self.workbook_bytes(
            headers=["rut", "fecha_inicio", "fecha_fin"],
            rows=[["20.728.268", "26-01-26", "30-01-26"]],
        )

        salida, _meta = modificar_archivo(contenido, ".xlsx", "Municipal")
        result = load_workbook(filename=BytesIO(salida.getvalue()))
        ws = result.active

        self.assertEqual(ws.cell(row=2, column=5).value, "NO")
        self.assertIsNone(ws.cell(row=2, column=6).value)


if __name__ == "__main__":
    unittest.main()
